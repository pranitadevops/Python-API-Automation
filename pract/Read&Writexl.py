import openpyxl

wb = openpyxl.load_workbook("C:/Users/ptalekar/Downloads/ReadXL.xlsx")
sh1 = wb['Automation']
row = sh1.max_row
column = sh1.max_column
for i in range(1, row + 1):  # Read file
    for j in range(1, column + 1):
        print(sh1.cell(i, j).value)

# write into New or existing file
sh1.cell(row=5, column=1, value='pytest')
sh1.cell(row=5, column=2, value='Framework')
sh1.cell(row=5, column=3, value='764764786')

wb.save("Report.xlsx")

print(sh1.cell(2, 1).value)  # Read file
