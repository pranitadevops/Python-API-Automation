from textwrap import indent

import requests
import openpyxl
import json
import Test_Case.test_login

from Test_Case.Environment import Auth, baseUrl


def test_Multipledata():
    # API URL
    UserEndPoint = "/api/information"
    headers = {'Content-Type': 'application/json', 'Authorization': Auth + Test_Case.test_login.Bearer_token}
    # Read Payload
    f = open('C:/Users/ptalekar/Downloads/Multidata.json')
    payload = json.loads(f.read())

    # Excel Code
    wb = openpyxl.load_workbook('C:/Users/ptalekar/Downloads/Multidata.xlsx')
    sh = wb['Sheet1']
    row = sh.max_row
    # Read Data
    for i in range(2, row + 1):
        cell_image = sh.cell(row=i, column=1)
        cell_data = sh.cell(row=i, column=2)
        cell_text = sh.cell(row=i, column=3)
        cell_title = sh.cell(row=i, column=4)
        cell_type = sh.cell(row=i, column=5)

        payload['image'] = cell_image.value
        payload['imageContent'][0] = cell_data.value
        payload['text'] = cell_text.value
        payload['title'] = cell_title.value
        payload['type'] = cell_type.value

        response = requests.post(baseUrl + UserEndPoint, json=payload, headers=headers)
        print(response.text)
        print(payload['type'])
        if payload['type'] != 'CHRONIC_KIDNEY_DISEASE':
            print("Unexpected Title")

    # if response.status_code >= 400:
    #     raise RuntimeError("Failed : HTTP error code : " + str(response.status_code))





















# Write into file
# sh.cell(row=5, column=1, value='pytest')
# sh.cell(row=5, column=2, value='Framework')
# sh.cell(row=5, column=3, value='Kabi_colombia')
# sh.cell(row=5, column=4, value='THE_KIDNEYS')
# sh.cell(row=5, column=5, value='string')
# wb.save('Multidata.xlsx')
# print(str(payload))
